"""
Modul Report Generator untuk Sistem Validasi RPS-BAP.

Modul ini menyusun dan mengekspor laporan hasil validasi
kesesuaian RPS-BAP dalam format PDF and Excel.

Sesuai PRD - Modul Report Generator.
"""

import os
from typing import List, Optional, Dict, Any
from datetime import datetime

from models.report import Report
from models.validation_result import ValidationResult
from repositories.rps_repository import RPSRepository
from repositories.bap_repository import BAPRepository
from repositories.validation_repository import ValidationRepository
from utils.logger import setup_logger
from utils.exceptions import ReportGenerationError
from config.constants import (
    STATUS_SESUAI,
    STATUS_TIDAK_SESUAI,
    STATUS_TIDAK_DITEMUKAN,
    DATETIME_FORMAT,
    MSG_REPORT_FAILED,
)

# Inisialisasi logger untuk modul ini
logger = setup_logger(__name__)


class ReportGenerator:
    """
    Class untuk membuat dan mengekspor laporan validasi.
    """

    def __init__(
        self,
        rps_repo: RPSRepository,
        bap_repo: BAPRepository,
        validation_repo: ValidationRepository
    ):
        """
        Inisialisasi ReportGenerator dengan dependency injection.

        Args:
            rps_repo: Repository untuk akses data RPS.
            bap_repo: Repository untuk akses data BAP.
            validation_repo: Repository untuk akses hasil validasi.
        """
        self._rps_repo: RPSRepository = rps_repo
        self._bap_repo: BAPRepository = bap_repo
        self._validation_repo: ValidationRepository = validation_repo
        logger.info("ReportGenerator berhasil diinisialisasi")

    def generate_report(self) -> Report:
        """
        Membuat laporan kesesuaian pembelajaran aktif.

        Returns:
            Report: Objek laporan yang berisi seluruh hasil analisis.

        Raises:
            ReportGenerationError: Jika data validasi belum tersedia.
        """
        logger.info("Membuat laporan kesesuaian pembelajaran")

        # Mengambil hasil validasi dari database
        validation_results = self._validation_repo.get_all()
        # Jika belum ada hasil validasi, kembalikan laporan kosong (bukan error)
        if not validation_results:
            logger.warning("Belum ada hasil validasi aktif — mengembalikan laporan kosong.")
            return Report(
                compliance_percentage=0.0,
                total_meetings=0,
                matched_count=0,
                mismatched_list=[],
                missing_list=[],
                results=[],
                generated_at=datetime.now()
            )

        # Mengambil data RPS dan BAP untuk konteks laporan
        rps_list = self._rps_repo.get_all()
        bap_list = self._bap_repo.get_all()

        # Menghitung statistik kesesuaian (BR-09)
        total_meetings = len(rps_list)
        sesuai_count = sum(
            1 for r in validation_results if r.status == STATUS_SESUAI
        )
        compliance_percentage = (
            (sesuai_count / total_meetings) * 100
            if total_meetings > 0 else 0.0
        )

        # Menyusun daftar materi tidak sesuai (AC-14)
        mismatched_list = self._build_mismatched_list(
            validation_results, rps_list, bap_list
        )

        # Menyusun daftar materi belum diajarkan (AC-15)
        bap_meetings = {bap.meeting_number for bap in bap_list}
        missing_list = [
            {
                "meeting_number": rps.meeting_number,
                "topic": rps.topic,
                "sub_topic": rps.sub_topic or "",
            }
            for rps in rps_list
            if rps.meeting_number not in bap_meetings
        ]

        # Menyusun objek Report dengan data RPS/BAP yang diperkaya
        rps_map = {rps.meeting_number: rps for rps in rps_list}
        bap_map = {bap.meeting_number: bap for bap in bap_list}
        enriched_results = []
        for r in validation_results:
            r_dict = r.to_dict()
            rps_item = rps_map.get(r.meeting_number)
            bap_item = bap_map.get(r.meeting_number)
            r_dict["rps_topic"] = (rps_item.topic or "") if rps_item else ""
            r_dict["bap_material"] = (bap_item.material_taught or "") if bap_item else ""
            enriched_results.append(r_dict)

        report = Report(
            compliance_percentage=round(compliance_percentage, 2),
            total_meetings=total_meetings,
            matched_count=sesuai_count,
            mismatched_list=mismatched_list,
            missing_list=missing_list,
            results=enriched_results,
            generated_at=datetime.now()
        )

        logger.info(
            f"Laporan berhasil dibuat. Kesesuaian: {compliance_percentage:.2f}%"
        )
        return report

    def _build_mismatched_list(
        self,
        validation_results: List[ValidationResult],
        rps_list: list,
        bap_list: list
    ) -> List[Dict[str, Any]]:
        """
        Menyusun daftar materi yang tidak sesuai atau tidak ditemukan.
        """
        rps_map = {rps.meeting_number: rps for rps in rps_list}
        bap_map = {bap.meeting_number: bap for bap in bap_list}

        mismatched = []
        for result in validation_results:
            if result.status in (STATUS_TIDAK_SESUAI, STATUS_TIDAK_DITEMUKAN):
                rps = rps_map.get(result.meeting_number)
                bap = bap_map.get(result.meeting_number)

                mismatched.append({
                    "meeting_number": result.meeting_number,
                    "status": result.status,
                    "similarity_score": result.similarity_score,
                    "rps_topic": rps.topic if rps else "-",
                    "bap_material": bap.material_taught if bap else "-",
                    "notes": result.notes or "",
                })

        return mismatched

    def export_to_pdf(self, report: Report, output_path: str) -> str:
        """
        Mengekspor laporan ke format PDF profesional menggunakan reportlab.
        """
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.units import cm
            from reportlab.platypus import (
                SimpleDocTemplate, Table, TableStyle,
                Paragraph, Spacer
            )
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

            logger.info(f"Mengekspor laporan ke PDF: {output_path}")

            os.makedirs(os.path.dirname(output_path), exist_ok=True)

            class NumberedDocTemplate(SimpleDocTemplate):
                def afterPage(self):
                    self.canv.saveState()
                    self.canv.setFont('Helvetica', 8)
                    self.canv.drawCentredString(
                        A4[0] / 2, 1 * cm,
                        f"Halaman {self.page}"
                    )
                    self.canv.drawString(
                        2 * cm, 1 * cm,
                        f"Laporan Validasi RPS-BAP | {datetime.now().strftime('%Y-%m-%d %H:%M')}"
                    )
                    self.canv.restoreState()

            doc = NumberedDocTemplate(
                output_path,
                pagesize=A4,
                rightMargin=2 * cm,
                leftMargin=2 * cm,
                topMargin=2 * cm,
                bottomMargin=2.5 * cm
            )

            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=14,
                leading=18,
                spaceAfter=6,
                alignment=1,
                fontName='Helvetica-Bold'
            )
            subtitle_style = ParagraphStyle(
                'CustomSubtitle',
                parent=styles['Heading2'],
                fontSize=11,
                leading=14,
                spaceAfter=6,
                spaceBefore=10,
                fontName='Helvetica-Bold'
            )
            normal_style = ParagraphStyle(
                'CustomNormal',
                parent=styles['Normal'],
                fontSize=10,
                leading=13,
                spaceAfter=4
            )
            small_style = ParagraphStyle(
                'CustomSmall',
                parent=styles['Normal'],
                fontSize=9,
                leading=11,
                spaceAfter=2
            )

            elements = []

            # Judul
            elements.append(Paragraph("LAPORAN VALIDASI KESESUAIAN RPS DAN BAP", title_style))
            elements.append(Spacer(1, 6))

            # Informasi laporan
            now_str = datetime.now().strftime("%d %B %Y %H:%M")
            info_lines = [
                f"Tanggal Cetak: {now_str}",
                f"Total Pertemuan RPS: {report.total_meetings}",
                f"Total Realisasi BAP: {report.matched_count + len(report.mismatched_list)}",
                f"Jumlah Sesuai: {report.matched_count}",
                f"Jumlah Tidak Sesuai / Tidak Ditemukan: {len(report.mismatched_list)}",
                f"Persentase Kesesuaian: {report.compliance_percentage:.2f}%",
            ]
            for line in info_lines:
                elements.append(Paragraph(line, normal_style))
            elements.append(Spacer(1, 12))

            # Tabel detail validasi
            elements.append(Paragraph("Detail Hasil Validasi", subtitle_style))

            detail_header = ["No", "Pertemuan", "Topik RPS", "Realisasi BAP", "Status", "Persentase", "Catatan"]
            detail_data = [detail_header]

            detail_col_widths = [1 * cm, 2 * cm, 5 * cm, 5 * cm, 2.5 * cm, 2 * cm, 4.5 * cm]

            header_style = ParagraphStyle(
                'HeaderCell', parent=normal_style,
                fontSize=9, leading=11, fontName='Helvetica-Bold',
                textColor=colors.white, alignment=1
            )
            cell_style = ParagraphStyle(
                'CellStyle', parent=small_style,
                fontSize=8, leading=10
            )
            cell_center = ParagraphStyle(
                'CellCenter', parent=cell_style,
                alignment=1
            )

            def p(text, style=cell_style):
                return Paragraph(str(text) if text else "-", style)

            for idx, r in enumerate(report.results, 1):
                meeting = r.get("meeting_number", "-")
                topic = r.get("rps_topic", "") or "-"
                material = r.get("bap_material", "") or "-"
                status = r.get("status", "-")
                score = r.get("similarity_score", 0)
                notes = r.get("notes", "") or "-"
                detail_data.append([
                    p(str(idx), cell_center),
                    p(f"Pert. {meeting}", cell_center),
                    p(topic),
                    p(material),
                    p(status, cell_center),
                    p(f"{score:.2f}%", cell_center),
                    p(notes[:100]),
                ])

            detail_table = Table(detail_data, colWidths=detail_col_widths, repeatRows=1)
            header_bg = colors.HexColor('#4472C4')
            alt_bg = colors.HexColor('#D6E4F0')
            detail_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), header_bg),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, alt_bg]),
            ]))
            elements.append(detail_table)

            doc.build(elements)
            logger.info(f"Laporan PDF berhasil diekspor ke {output_path}")
            return os.path.abspath(output_path)

        except ImportError as e:
            logger.error(f"Library reportlab tidak tersedia: {e}")
            raise ReportGenerationError(
                "Library reportlab tidak terinstall",
                details={"error": str(e)}
            )
        except Exception as e:
            logger.exception(f"Gagal mengekspor laporan ke PDF: {e}")
            raise ReportGenerationError(
                MSG_REPORT_FAILED,
                details={"format": "PDF", "error": str(e)}
            )

    def export_to_excel(self, report: Report, output_path: str) -> str:
        """
        Mengekspor laporan ke format Excel profesional menggunakan openpyxl.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import (
                Font, PatternFill, Alignment, Border, Side
            )
            from openpyxl.utils import get_column_letter

            logger.info(f"Mengekspor laporan ke Excel: {output_path}")

            os.makedirs(os.path.dirname(output_path), exist_ok=True)

            wb = Workbook()
            ws = wb.active
            ws.title = "Laporan Validasi"

            # Style definitions
            blue_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            white_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
            title_font = Font(name="Calibri", bold=True, size=14)
            info_font = Font(name="Calibri", size=11)
            header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
            cell_font = Font(name="Calibri", size=10)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

            row = 1
            # Title
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
            title_cell = ws.cell(row=row, column=1, value="LAPORAN VALIDASI KESESUAIAN RPS DAN BAP")
            title_cell.font = title_font
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 2

            # Info
            now_str = datetime.now().strftime("%d %B %Y %H:%M")
            info_data = [
                ("Tanggal Cetak:", now_str),
                ("Total Pertemuan RPS:", str(report.total_meetings)),
                ("Total Realisasi BAP:", str(report.matched_count + len(report.mismatched_list))),
                ("Jumlah Sesuai:", str(report.matched_count)),
                ("Jumlah Tidak Sesuai / Tidak Ditemukan:", str(len(report.mismatched_list))),
                ("Persentase Kesesuaian:", f"{report.compliance_percentage:.2f}%"),
            ]
            for label, value in info_data:
                c1 = ws.cell(row=row, column=1, value=label)
                c1.font = Font(name="Calibri", bold=True, size=10)
                c2 = ws.cell(row=row, column=2, value=value)
                c2.font = info_font
                row += 1
            row += 1

            # Table header
            headers = ["No", "Pertemuan", "Topik RPS", "Realisasi BAP", "Status", "Persentase", "Catatan"]
            for col_idx, h in enumerate(headers, 1):
                c = ws.cell(row=row, column=col_idx, value=h)
                c.font = header_font
                c.fill = blue_fill
                c.alignment = center_align
                c.border = thin_border
            header_row = row
            row += 1

            # Table data
            alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
            for idx, r in enumerate(report.results, 1):
                data_row = [
                    idx,
                    f"Pertemuan {r.get('meeting_number', '-')}",
                    r.get("rps_topic", "") or "-",
                    r.get("bap_material", "") or "-",
                    r.get("status", "-"),
                    f"{r.get('similarity_score', 0):.2f}%",
                    r.get("notes", "") or "-",
                ]
                for col_idx, val in enumerate(data_row, 1):
                    c = ws.cell(row=row, column=col_idx, value=val)
                    c.font = cell_font
                    c.border = thin_border
                    if col_idx in (1, 5, 6):
                        c.alignment = center_align
                    else:
                        c.alignment = left_align
                    if (idx % 2) == 0:
                        c.fill = alt_fill
                row += 1

            data_end_row = row - 1

            # Column widths
            col_widths = [5, 12, 30, 30, 15, 12, 35]
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

            # Freeze header
            ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
            ws.auto_filter.ref = f"A{header_row}:G{data_end_row}"

            wb.save(output_path)
            logger.info(f"Laporan Excel berhasil diekspor ke {output_path}")
            return os.path.abspath(output_path)

        except ImportError as e:
            logger.error(f"Library openpyxl tidak tersedia: {e}")
            raise ReportGenerationError(
                "Library openpyxl tidak terinstall",
                details={"error": str(e)}
            )
        except Exception as e:
            logger.exception(f"Gagal mengekspor laporan ke Excel: {e}")
            raise ReportGenerationError(
                MSG_REPORT_FAILED,
                details={"format": "Excel", "error": str(e)}
            )
